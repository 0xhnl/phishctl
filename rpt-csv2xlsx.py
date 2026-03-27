#!/usr/bin/env python3
import pandas as pd
import argparse
import glob
import os
import re
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

def style_sheet(ws):
    header_fill = PatternFill(start_color="2F2F2F", end_color="2F2F2F", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    center_align = Alignment(horizontal="center", vertical="center")

    # Style header row
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center_align

    # Auto column width
    for col in ws.columns:
        max_length = 0
        col_letter = get_column_letter(col[0].column)

        for cell in col:
            try:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            except:
                pass

        ws.column_dimensions[col_letter].width = max_length + 2


def combine_csv_to_excel(input_folder, output_file):
    all_files = glob.glob(os.path.join(input_folder, "*.csv"))
    if not all_files:
        print("No CSV files found.")
        return

    pattern = re.compile(r'^([CF]\d+(?:-\d+)?)_results\.csv$')
    campaigns = []

    for file in all_files:
        filename = os.path.basename(file)
        match = pattern.match(filename)
        if match:
            campaign_name = match.group(1)
            campaigns.append((campaign_name, file))

    if not campaigns:
        print("No valid campaign files.")
        return

    def sort_key(name):
        match = re.match(r'^([CF])(\d+)(?:-(\d+))?$', name)
        main = int(match.group(2))
        sub = int(match.group(3)) if match.group(3) is not None else -1
        return (match.group(1), main, sub)

    campaigns.sort(key=lambda item: sort_key(item[0]))
    sorted_keys = [name for name, _ in campaigns]

    summary_results = {}
    dataframes = {}

    # Read all first
    for key, file_path in campaigns:
        df = pd.read_csv(file_path)

        desired_columns = [
            'id', 'email', 'status',
            'ip', 'latitude', 'longitude', 'send_date'
        ]

        df = df[[col for col in desired_columns if col in df.columns]]

        dataframes[key] = df

        total = len(df)
        opened = len(df[df['status'].str.contains('open', case=False, na=False)])
        clicked = len(df[df['status'].str.contains('click', case=False, na=False)])
        submitted = len(df[df['status'].str.contains('submit', case=False, na=False)])
        ignored = total - opened - clicked - submitted

        summary_results[key] = [
            opened, clicked, submitted, ignored, total
        ]

    # Create summary dataframe
    summary_index = [
        "Email Opened",
        "Links Clicked",
        "Submitted Data (Fell for trap)",
        "Email Ignored (No Action)",
        "Total Employee"
    ]

    summary_df = pd.DataFrame(summary_results, index=summary_index)

    # Write Excel
    with pd.ExcelWriter(output_file, engine='openpyxl') as writer:

        # 1️⃣ Write Summary FIRST
        summary_df.to_excel(writer, sheet_name="Summary")

        # 2️⃣ Write campaign sheets
        for key in sorted_keys:
            dataframes[key].to_excel(writer, sheet_name=key, index=False)

        # Apply styling
        workbook = writer.book

        # Style Summary
        ws_summary = workbook["Summary"]
        style_sheet(ws_summary)

        # Light highlight first column
        light_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
        for cell in ws_summary["A"]:
            cell.fill = light_fill

        # Style all campaign sheets
        for key in sorted_keys:
            ws = workbook[key]
            style_sheet(ws)

    print(f"\n✅ Professional report generated: {output_file}")


if __name__ == "__main__":
    parser = argparse.ArgumentParser(description="Professional Campaign Report Generator")
    parser.add_argument("-ff", "--folder", required=True, help="CSV folder")
    parser.add_argument("-o", "--output", required=True, help="Output Excel file")

    args = parser.parse_args()

    combine_csv_to_excel(args.folder, args.output)
